import re
import openpyxl
from decimal import Decimal
from django.db import transaction
from .models import ImportacionPagos, PagoAlumno

class ImportadorPagosExcel:
    
    def __init__(self, archivo, anio, usuario=None):
        self.archivo = archivo
        self.anio = anio
        self.usuario = usuario
        self.errores = []
        self.importados = 0
        
    def limpiar_datos_existentes(self):
        """Elimina todos los pagos del año especificado"""
        deleted_count = PagoAlumno.objects.filter(importacion__anio=self.anio).delete()
        ImportacionPagos.objects.filter(anio=self.anio).delete()
        # print(f"Eliminados {deleted_count[0] if deleted_count else 0} registros")
        return deleted_count[0] if deleted_count else 0
    
    def limpiar_valor(self, valor):
        """Limpia y devuelve el valor tal cual del Excel"""
        if valor is None or valor == '':
            return '-'
        return str(valor).strip()
    
    def extraer_dni_de_doc_facturacion(self, doc_facturacion):
        """
        Extrae el DNI de la cadena de Doc. Facturación
        Ejemplo: "DNI: 62803376" -> "62803376"
        """
        if not doc_facturacion or doc_facturacion == '-':
            return ''
        
        # Buscar patrón DNI: XXXXXXXX o DNI:XXXXXXXX
        patrones = [
            r'DNI:\s*(\d{8})',      # DNI: 12345678 o DNI:12345678
            r'DNI\s*:\s*(\d{8})',    # DNI : 12345678
            r'(\d{8})',              # Solo 8 dígitos (como respaldo)
        ]
        
        for patron in patrones:
            match = re.search(patron, str(doc_facturacion))
            if match:
                return match.group(1)
        
        return ''
    
    def es_fila_valida(self, fila):
        """Verifica si la fila tiene al menos la columna Estudiante (columna 2)"""
        if len(fila) < 2:
            return False
        
        estudiante = fila[1] if len(fila) > 1 else ''
        estudiante = str(estudiante).strip() if estudiante else ''
        
        if not estudiante or estudiante == 'None' or estudiante == '-':
            return False
        
        # Excluir filas que son encabezados
        encabezados = ['estudiante', 'dni', 'total', 'pagado', 'reporte', 'clientes']
        if estudiante.lower() in encabezados:
            return False
        
        return True
    
    def procesar_fila(self, fila, importacion, row_num):
        """Procesa una fila del Excel y crea un registro horizontal"""
        
        if len(fila) < 2:
            return None
        
        # Extraer datos básicos
        num = fila[0] if len(fila) > 0 and fila[0] else None
        estudiante = self.limpiar_valor(fila[1]) if len(fila) > 1 else ''
        
        # DNI de la columna original (podría estar vacío)
        dni_original = ''
        if len(fila) > 2 and fila[2]:
            dni_original = self.limpiar_valor(fila[2])
        
        # Doc. Facturación (de aquí extraeremos el DNI si está disponible)
        doc_facturacion = self.limpiar_valor(fila[3]) if len(fila) > 3 else ''
        
        # Extraer DNI de Doc. Facturación
        dni_extraido = self.extraer_dni_de_doc_facturacion(doc_facturacion)
        
        # Priorizar: usar DNI extraído de Doc. Facturación, si no, usar el DNI original
        dni_final = dni_extraido if dni_extraido else dni_original
        
        nombre_facturacion = self.limpiar_valor(fila[4]) if len(fila) > 4 else ''
        nivel = self.limpiar_valor(fila[5]) if len(fila) > 5 else ''
        grado = self.limpiar_valor(fila[6]) if len(fila) > 6 else ''
        seccion = self.limpiar_valor(fila[7]) if len(fila) > 7 else ''
        
        # Pagos por mes (columnas 8 a 17)
        marzo = self.limpiar_valor(fila[8]) if len(fila) > 8 else '-'
        abril = self.limpiar_valor(fila[9]) if len(fila) > 9 else '-'
        mayo = self.limpiar_valor(fila[10]) if len(fila) > 10 else '-'
        junio = self.limpiar_valor(fila[11]) if len(fila) > 11 else '-'
        julio = self.limpiar_valor(fila[12]) if len(fila) > 12 else '-'
        agosto = self.limpiar_valor(fila[13]) if len(fila) > 13 else '-'
        setiembre = self.limpiar_valor(fila[14]) if len(fila) > 14 else '-'
        octubre = self.limpiar_valor(fila[15]) if len(fila) > 15 else '-'
        noviembre = self.limpiar_valor(fila[16]) if len(fila) > 16 else '-'
        diciembre = self.limpiar_valor(fila[17]) if len(fila) > 17 else '-'
        
        # Totales
        total = self.limpiar_valor(fila[18]) if len(fila) > 18 else '0'
        pagado = self.limpiar_valor(fila[19]) if len(fila) > 19 else '0'
        
        # Debug: mostrar cómo se extrae el DNI
        # if row_num <= 10:  # Solo mostrar primeras 10 filas
        #     print(f"Fila {row_num}:")
        #     print(f"  - Doc. Facturación: {doc_facturacion}")
        #     print(f"  - DNI extraído: {dni_extraido}")
        #     print(f"  - DNI original: {dni_original}")
        #     print(f"  - DNI final: {dni_final}")
        #     print(f"  - Estudiante: {estudiante}")
        #     print("-" * 50)
        
        # Crear registro
        pago = PagoAlumno(
            num=num,
            estudiante=estudiante,
            dni=dni_final,  # Usar el DNI extraído o el original
            doc_facturacion=doc_facturacion,
            nombre_facturacion=nombre_facturacion,
            nivel=nivel,
            grado=grado,
            seccion=seccion,
            marzo=marzo,
            abril=abril,
            mayo=mayo,
            junio=junio,
            julio=julio,
            agosto=agosto,
            setiembre=setiembre,
            octubre=octubre,
            noviembre=noviembre,
            diciembre=diciembre,
            total=total,
            pagado=pagado,
            importacion=importacion
        )
        return pago
    
    @transaction.atomic
    def importar(self, limpiar_antes=True):
        """Método principal de importación"""
        
        # if limpiar_antes:
        #     eliminados = self.limpiar_datos_existentes()
        #     # print(f"✅ Eliminados {eliminados} registros anteriores del año {self.anio}")
        
        # Crear registro de importación
        importacion = ImportacionPagos.objects.create(
            nombre_archivo=self.archivo.name,
            usuario=self.usuario or 'system',
            anio=self.anio
        )
        
        try:
            # Cargar el archivo Excel
            workbook = openpyxl.load_workbook(self.archivo, data_only=True)
            sheet = workbook.active
            
            # print(f"\n📊 PROCESANDO ARCHIVO EXCEL")
            # print(f"📄 Nombre: {self.archivo.name}")
            # print(f"📏 Total filas en Excel: {sheet.max_row}")
            # print(f"📏 Total columnas: {sheet.max_column}")
            
            registros_creados = []
            filas_procesadas = 0
            filas_invalidas = 0
            
            # Buscar la primera fila con datos
            inicio_datos = 1
            for row_num in range(1, min(20, sheet.max_row + 1)):
                fila = [cell.value for cell in sheet[row_num]]
                if self.es_fila_valida(fila):
                    inicio_datos = row_num
                    # print(f"✅ Primera fila con datos encontrada en fila {row_num}")
                    break
            
            # print(f"📍 Iniciando importación desde fila {inicio_datos}")
            # print(f"\n🔍 Extrayendo DNIs de la columna 'Doc. Facturación'...\n")
            
            # Empezar desde la fila donde se encontraron los datos
            for row_num in range(inicio_datos, sheet.max_row + 1):
                fila = [cell.value for cell in sheet[row_num]]
                
                if self.es_fila_valida(fila):
                    try:
                        pago = self.procesar_fila(fila, importacion, row_num)
                        if pago:
                            registros_creados.append(pago)
                            self.importados += 1
                            filas_procesadas += 1
                            
                            # if self.importados % 50 == 0:
                            #     print(f"   Procesados {self.importados} registros...")
                                
                    except Exception as e:
                        error_msg = f"Error en fila {row_num}: {str(e)}"
                        self.errores.append(error_msg)
                        # print(f"❌ {error_msg}")
                else:
                    filas_invalidas += 1
                    if filas_invalidas > 20 and self.importados == 0:
                        # print(f"⚠️ No se encontraron filas válidas con estudiantes")
                        break
            
            # Crear todos los registros en batch
            if registros_creados:
                PagoAlumno.objects.bulk_create(registros_creados)
            
            # Estadísticas de DNIs extraídos
            # dnis_extraidos = sum(1 for p in registros_creados if p.dni and len(p.dni) >= 8)
            # dnis_vacios = sum(1 for p in registros_creados if not p.dni)
            
            # print(f"\n📊 RESUMEN DE IMPORTACIÓN:")
            # print(f"   ✅ Registros importados: {self.importados}")
            # print(f"   📇 DNIs extraídos de Doc. Facturación: {dnis_extraidos}")
            # print(f"   ⚠️  DNIs vacíos: {dnis_vacios}")
            # print(f"   ⚠️  Filas sin estudiante: {filas_invalidas}")
            # print(f"   📝 Errores: {len(self.errores)}")
            
            # Actualizar el total de registros en la importación
            importacion.total_registros = self.importados
            importacion.save()
            
            return {
                'success': True,
                'importados': self.importados,
                'errores': self.errores,
                'importacion_id': importacion.id
            }
            
        except Exception as e:
            error_msg = f"Error general al procesar el archivo: {str(e)}"
            self.errores.append(error_msg)
            # print(f"❌ {error_msg}")
            import traceback
            traceback.print_exc()
            return {
                'success': False,
                'importados': 0,
                'errores': self.errores
            }